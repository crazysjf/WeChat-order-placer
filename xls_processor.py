# -*- coding: utf-8 -*-
from openpyxl import load_workbook
import openpyxl.styles as sty
from openpyxl.styles import Border, Side, Font, Alignment

import utils
import re
import constants
import pandas as pd

class XlsProcessor():
    # 修改后的填充色
    MODIFICARTION_FILL = sty.PatternFill(fill_type='solid', fgColor=constants.XLS_FG_COLOR_WARNING)

    def __init__(self, f):
        self._f = f

    # def _get_xls_column_nr(self, table_head, name):
    #     for i in range(0, len(table_head)):
    #         if table_head[i] == name:
    #             return i
    #     return None

    def _open(self):
        self.wb = load_workbook(self._f)
        self.ws = self.wb.active
        self.nrows = self.ws.max_row
        self.good_code_cn = self._get_column_cn(u'商品编码')
        self.provider_cn = self._get_column_cn(u'供应商')
        self.code_cn = self._get_column_cn(u'供应商款号')
        self.spec_cn = self._get_column_cn(u'颜色规格')
        self.nr_cn = self._get_column_cn(u'数量')
        self.payed_cn = self._get_column_cn("实付")
        self.received_cn = self._get_column_cn("实拿")
        self.notation_cn = self._get_column_cn("商品备注")
        self.cost_cn = self._get_column_cn("成本价")
        self.sum_cn = self._get_column_cn("金额")


    def _close(self):
        self.wb.close()
        self.wb = None
        self.ws = None
        self.nrows = 0

    def _save(self):
        while (True):
            try:
                self.wb.save(self._f)
                break
            except IOError:
                str = input("Save failed. File may be open in other application. Retry? (Y/n)")
                if str == 'n' or str == 'N':
                    break


    def _get_column_cn(self, name):
        for c in range(1, self.ws.max_column + 1):
            v = self.ws.cell(row = 1, column=c).value
            if v == name:
                return c
        return None



    def gen_orders(self):
        '''
        生成所有订单并返回，格式：
        {
        档口1: [{code: 商品编码1， spec: 规格1, nr: 数量1}，
                {code: 商品编码2， spec: 规格2, nr: 数量2，... ],
        档口2: [{code: 商品编码1， spec: 规格1, nr: 数量1}，
                {code: 商品编码2， spec: 规格2, nr: 数量2，... ],
        ...
        }
        '''
        # wb = load_workbook(self._f)
        # ws = wb.active
        # nrows = ws.max_row

        self._open()

        provider_cn = self._get_column_cn( u'供应商')
        code_cn = self._get_column_cn(u'供应商款号')
        spec_cn = self._get_column_cn(u'颜色规格')
        nr_cn = self._get_column_cn(u'数量')

        orders = {}  # 解析之后的所有订单，键值为档口名
        provider_order = []  # 单个档口的订单

        for i in range(2, self.nrows):
            provider = utils.convert_possible_num_to_str(self.ws.cell(row=i, column=provider_cn).value)
            code = utils.convert_possible_num_to_str(self.ws.cell(row=i, column=code_cn).value)

            if provider == "**" or provider == u"样衣":
                # 截止到内容未**，或者“样衣”两个字的行
                break

            # 跳过空行和汇总行。
            if provider == None or \
                            provider.find(u'汇总') != -1 or \
                            provider.find(u'总计') != -1:
                continue
            #print provider, code

            order_line = {}
            order_line['code'] = code
            order_line['spec'] = self.ws.cell(row=i, column=spec_cn).value
            order_line['nr'] = utils.convert_possible_num_to_str(self.ws.cell(row=i, column=nr_cn).value)

            # 如果是第一次换次品，则次品不报单，避免次品没有送过去，档口看不懂报表
            # 如果当天没有换回，第二天会变成欠货，则会报单
            if ('次品' in str(order_line['spec'])) and ('换' in order_line['nr']):
                pass
            else:
                provider_order.append(order_line)

                if provider in orders:
                    orders[provider].append(order_line)
                else:
                    orders[provider] = [order_line]

        self._close()
        return orders

    def annotate_unknown_providers(self, unknown_provider):
        '''
        把未知供应商标记为红色。
        成功返回True
        如果文件被其他应用打开，标注失败，返回False。
        :param unknown_provider: 
        :return: 
        '''
        self._open()

        provider_cn = self._get_column_cn(u'供应商')
        code_cn = self._get_column_cn(u'供应商款号')

        # 写入数据
        for i in range(2, self.nrows):
            code = utils.convert_possible_num_to_str(self.ws.cell(row=i, column=code_cn).value)
            # 跳过汇总行
            if code == None:
                continue

            cell = self.ws.cell(row=i, column=provider_cn)
            v = cell.value
            for p in unknown_provider:
                if v == p:
                    cell.fill = sty.PatternFill(fill_type='solid', fgColor="ff6347")

        self._save()
        self._close()
        return True


    def calc_order_exceptions(self):
        '''获取报单异常，包括：
        欠货，
        其他
        返回： 格式：
        {
        档口1: [{code: 商品编码1， spec: 规格1, nr: 欠货数1,  'received':到货数1, 'notation': 备注, }，
                {code: 商品编码2， spec: 规格2, nr: 欠货数2, 'received':到货数2... ],
        档口2: [{code: 商品编码1， spec: 规格1, nr: 欠货数1,'received':到货数1, 'notation':备注},
                {code: 商品编码2， spec: 规格2, nr: 欠货数2, 'received':到货数2... ],
        ...
        }
        如果有'notation'键，表明备注需要发送给档口
        '''
        # 处理欠，报，换等几种情况
        self._open()
        # provider_cn = self._get_column_cn(u'供应商')
        # code_cn = self._get_column_cn(u'供应商款号')
        # spec_cn = self._get_column_cn(u'颜色规格')
        # nr_cn = self._get_column_cn(u'数量')
        # payed_cn = self._get_column_cn("实付")
        # received_cn = self._get_column_cn("实拿")

        result = {}

        for i in range(2, self.nrows):
            provider = utils.convert_possible_num_to_str(self.ws.cell(row=i, column=self.provider_cn).value)
            code = utils.convert_possible_num_to_str(self.ws.cell(row=i, column=self.code_cn).value)
            spec = self.ws.cell(row=i, column=self.spec_cn).value

            # 截止到内容未**，或者“样衣”两个字的行
            if provider == "**" or provider == u"样衣":
                break

            # 跳过空行和汇总行。
            if provider == None or \
                            provider.find(u'汇总') != -1 or \
                            provider.find(u'总计') != -1:
                continue

            nr =  utils.convert_possible_num_to_str(self.ws.cell(row=i, column=self.nr_cn).value)
            payed_s = utils.convert_possible_num_to_str(self.ws.cell(row=i, column=self.payed_cn).value)
            received_s = utils.convert_possible_num_to_str(self.ws.cell(row=i, column=self.received_cn).value)
            abnormal_cn = self.sum_cn + 1 # 异常列。该列如果不为空，表示有异常，备注需要发送给档口
            abnormal = utils.convert_possible_num_to_str(self.ws.cell(row=i, column=abnormal_cn).value)
            notation = utils.convert_possible_num_to_str(self.ws.cell(row=i, column=self.notation_cn).value)
            received, e = utils.calc_received_exceptions(nr, payed_s, received_s)

            # if abnormal != None:
            #     print('标记异常', i, notation)

            if e != 0 or abnormal != None:
                line = {'code':code, 'spec':spec, 'nr': e, 'received':received}
                if abnormal != None:
                    if notation == None:
                        #print("%s, %s, %s: 可能出错误，备注需要作为异常采用，但是为空"%(provider,code,spec))
                        notation = "无异常，检查报表"
                    line['notation'] = notation
                if not provider in result:
                    result[provider] = [line]
                else:
                    result[provider].append(line)
        self._close()
        return result

    def _do_insertion(self, line_nr, provider, line):
        line_nr = line_nr + 1
        self.ws.insert_rows(line_nr)
        cell_p = self.ws.cell(row=line_nr, column=self.provider_cn)
        cell_c = self.ws.cell(row=line_nr, column=self.code_cn)
        cell_s = self.ws.cell(row=line_nr, column=self.spec_cn)
        cell_n = self.ws.cell(row=line_nr, column=self.nr_cn)

        cell_p.value = provider
        cell_c.value = line['code']
        cell_s.value = line['spec']
        cell_n.value = utils.gen_text_for_one_exception_line(line, simplified=True)


        cell_p.fill = self.MODIFICARTION_FILL
        cell_c.fill = self.MODIFICARTION_FILL
        cell_s.fill = self.MODIFICARTION_FILL
        cell_n.fill = self.MODIFICARTION_FILL



    def _insert_one_line(self, provider, line):
        '''
        算法：正确的做法是根据报单表生成一组数据，然后根据异常修改数据，再生成结果报表。
        此处为了兼容之前的做法，直接对报表进行修改。
        先遍历一边报表，根据供应商、商品编码、颜色尺码的相等情况，把行号记录到三个数组中：
        same_p: 供应商相等的所有行
        same_p_c:供应商、商品编码都相等的所有行
        same_p_c_s:供应商、商品编码、颜色尺码都相等的行
        
        然后根据这三个数组来计算异常应该插入的位置。
        :param provider: 
        :param line: 异常行
        :return: 
        '''
        code = line['code']
        spec = line['spec']

        same_p = []
        same_p_c = []
        same_p_c_s = []

        for i in range(2, self.nrows):
            p = utils.convert_possible_num_to_str(self.ws.cell(row=i, column=self.provider_cn).value)
            c = utils.convert_possible_num_to_str(self.ws.cell(row=i, column=self.code_cn).value)
            s = utils.convert_possible_num_to_str(self.ws.cell(row=i, column=self.spec_cn).value)

            # 截止到内容未**，或者“样衣”两个字的行
            if p == "**" or p == u"样衣":
                break

            # 跳过空行和汇总行。
            if p == None or \
                            p.find(u'汇总') != -1 or \
                            p.find(u'总计') != -1:
                continue

            if provider == p:
                same_p.append(i)
                if code == c:
                    same_p_c.append(i)
                    if spec == s:
                        same_p_c_s.append(i)

        #print(same_p, same_p_c, same_p_c_s)
        if len(same_p_c_s) != 0:
            # 供应商，款号，编码完全相同，直接修改
            line_nr = same_p_c_s[-1]
            cell = self.ws.cell(row=line_nr, column=self.nr_cn)

            # 如果异常只是简单欠货，而且欠货数量刚好等于报单数量，数量栏直接写成“欠xx”
            # 如果有其他异常，或者数量不等，则写成“欠xx，原始值”
            if not 'notation' in line.keys() and line['nr'] == cell.value:
                val = utils.gen_text_for_one_exception_line(line, simplified=True)
            else:
                orig_val = utils.convert_possible_num_to_str(cell.value)
                val = utils.gen_text_for_one_exception_line(line, simplified=True) + ',' + orig_val

            cell.value = val
            cell.fill = self.MODIFICARTION_FILL

        elif len(same_p_c) != 0:
            # 仅供应商，款号相同，在同款号后面插入空行
            line_nr = same_p_c[-1]
            self._do_insertion(line_nr, provider, line)

        elif len(same_p) != 0:
            # 仅供应商相同，在供应商最后一行插入空行
            line_nr = same_p[-1]
            self._do_insertion(line_nr, provider, line)

        else:
            # 连应商都找不到，寻找合适位置插入空行
            # for i in range(2, self.nrows):
            #     p = utils.convert_possible_num_to_str(self.ws.cell(row=i, column=self.provider_cn).value)
            #     c = utils.convert_possible_num_to_str(self.ws.cell(row=i, column=self.code_cn).value)
            #     s = utils.convert_possible_num_to_str(self.ws.cell(row=i, column=self.spec_cn).value)
            #
            #     # 截止到内容未**，或者“样衣”两个字的行
            #     if p == "**" or p == u"样衣":
            #         break
            #
            #     # 跳过空行和汇总行。
            #     if p == None or \
            #                     p.find(u'汇总') != -1 or \
            #                     p.find(u'总计') != -1:
            #         continue
            #
            #     _p = provider
            #     if i == 2 and _p < p:
            #         break
            #
            #     # 找到下一个非空行
            #     j = i + 1
            #     while(j < self.nrows):
            #         p2 = utils.convert_possible_num_to_str(self.ws.cell(row=j, column=self.provider_cn).value)
            #         if p2 == None:
            #             break
            #         j = j + 1
            #
            #     if p2 != None and p < _p and _p < p2:
            #         break

            # 目前字符串比较有问题，直接插入到最上方
            i = 1
            self._do_insertion(i, provider, line)


    def insert_exceptions(self, exceptions):
        self._open()

        provider_cn = self._get_column_cn(u'供应商')
        code_cn = self._get_column_cn(u'供应商款号')
        #self.ws.insert_rows(100)

        for p in exceptions:
            for l in exceptions[p]:
                self._insert_one_line(p, l)


        # # 写入数据
        try:
            self._save()
        except IOError:
            return False

        self._close()
        return True


    def _adjust_column_width(self, name, width):
        """调整列宽"""
        col = self._get_column_cn(name)
        self.ws.column_dimensions[utils.num_to_alphabet(col - 1)].width = width

    def format(self):
        self._open()

        # 分隔供应商，需要从下往上遍历
        for i in range(self.nrows, 2, -1):
            p1 = self.ws.cell(row=i, column=self.provider_cn).value
            p2 = self.ws.cell(row=i-1, column=self.provider_cn).value
            if p1 is None:
                continue

            if p1 != p2:
                self.ws.insert_rows(i)

        # TODO: 插入行之后self.nrows不能再用

        # 算账
        nr_cn = self._get_column_cn('数量')
        payed_cn = self._get_column_cn('实付')
        cost_cn = self._get_column_cn('成本价')

        amount_cn = cost_cn + 1
        self.ws.cell(row=1, column=amount_cn).value = '金额'

        # 公式模型：=IF(ISBLANK(L2),K2,IF(OR(L2="x",L2="X"),0, IF(ISNUMBER(L2), L2,K2))) *Q2
        # K：数量
        # L：实付
        # Q：成本价
        K = utils.num_to_alphabet(nr_cn - 1)
        L = utils.num_to_alphabet(payed_cn - 1)
        Q = utils.num_to_alphabet(cost_cn - 1)

        for i in range(2, self.ws.max_row + 1):
            if self.ws.cell(row=i, column=self.provider_cn).value is None and \
                    self.ws.cell(row=i, column=self.code_cn).value is None:
                continue

            # 以下公式实，付栏为空当成照单付款
            #f = '=IF(ISBLANK({L}{ln}),{K}{ln},IF(OR({L}{ln}="x",{L}{ln}="X"),0, IF(ISNUMBER({L}{ln}), {L}{ln},{K}{ln}))) *{Q}{ln}'.format(K=K, L=L, Q=Q, ln=i)

            # 以下公式实，付栏为空当成未付款
            f = '=IF(ISBLANK({L}{ln}), 0,IF(OR({L}{ln}="x",{L}{ln}="X"),0, IF(ISNUMBER({L}{ln}), {L}{ln},{K}{ln}))) *{Q}{ln}'.format(K=K, L=L, Q=Q, ln=i)

            self.ws.cell(row=i, column=amount_cn).value = f

        # 交替填充款号背景
        gray_fill = sty.PatternFill(fill_type="solid", fgColor="F0F0F0")
        white_fill = sty.PatternFill(fill_type="solid", fgColor="FFFFFF")
        yellow_fill = sty.PatternFill(fill_type="solid", fgColor="FFEB9C")

        # old_p = self.ws.cell(row=2, column=self.provider_cn).value
        # old_c = self.ws.cell(row=2, column=self.code_cn).value

        fill = white_fill
        for i in range(3, self.ws.max_row + 1):
            p1 = self.ws.cell(row=i-1, column=self.provider_cn).value
            c1 = self.ws.cell(row=i-1, column=self.code_cn).value
            p2 = self.ws.cell(row=i, column=self.provider_cn).value
            c2 = self.ws.cell(row=i, column=self.code_cn).value
            if p1 != p2 or c1 != c2: # 检测到供应商或者编码变化
                fill = gray_fill if fill == white_fill else white_fill # 交换填充颜色

            if (p1 is None and c1 is None) or \
                    (p2 is None and c2 is None):
                fill = white_fill # 遇到空行重置为白底
                continue

            if fill == gray_fill:
                for cell in self.ws[i:i]:
                    cell.fill = gray_fill

        # 冻结首行
        self.ws.freeze_panes = "A2"

        # 整体风格：边框、字体
        thin = Side(border_style="thin", color="000000")
        font = Font(size=10)
        for r in self.ws[1:self.ws.max_row]:
            for cell in r:
                #print(cell.value)
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                cell.font = font

        # 局部风格
        notation_small_font = Font(size=7)
        day2_fill = sty.PatternFill(fill_type="solid", fgColor="FFCC33")
        day3_fill = sty.PatternFill(fill_type="solid", fgColor="FF003C")

        day_cn = self._get_column_cn('天数')
        nr_cn = self._get_column_cn('数量')
        suggested_nr_cn = self._get_column_cn('建议采购数')
        #cost_cn = self._get_column_cn('成本价')
        #self.ws.column_dimensions[chr(ord('A') + self.notation_cn - 1)].width = "20"

        self._adjust_column_width("款式编码", 15)
        self._adjust_column_width("商品编码", 18)
        self._adjust_column_width("供应商", 15)
        self._adjust_column_width("颜色规格", 15)

        self._adjust_column_width("商品备注", 20)
        self._adjust_column_width("天数", 3)
        self._adjust_column_width("上限天数", 3)
        self._adjust_column_width("成本价",4)
        self._adjust_column_width("金额", 6)

        for i in range(2, self.ws.max_row + 1):
            # 备注
            cell = self.ws.cell(row=i, column=self.notation_cn)
            notation = str(cell.value) if cell.value is not None else ""
            if '\n' in notation:
                cell.font = notation_small_font
                cell.alignment = Alignment(vertical='center',wrapText=True)

            if '收' in notation or "清" in notation or "销低" in notation:
                cell.fill = yellow_fill

            # 次品
            cell = self.ws.cell(row=i, column=self.code_cn)
            code = str(cell.value)
            if '次' in code:
                cell.fill = yellow_fill

            cell = self.ws.cell(row=i, column=self.spec_cn)
            spec = str(cell.value)
            if '次' in spec:
                cell.fill = yellow_fill

            # 天数
            cell = self.ws.cell(row=i, column=day_cn)
            s = cell.value
            if s is not None:
                d =  int(s[1:])
                if d == 2:
                    cell.fill = day2_fill
                elif d > 2:
                    cell.fill = day3_fill

            # 有异常的填红色背景
            cell = self.ws.cell(row=i, column=nr_cn)
            if cell.value is not None:
                s = str(cell.value)
                m = re.search(r'[^0-9]+', s)
                if m != None:
                    cell.fill = self.MODIFICARTION_FILL

            # 成本价大于阈值，且销小于3的，成本价高亮
            cell = self.ws.cell(row=i, column=cost_cn)
            s =  cell.value
            cost = int(s) if s is not None else 0

            s = self.ws.cell(row=i, column=suggested_nr_cn).value
            suggested_nr = int(s) if s is not None else 0

            if cost >= constants.COST_PRICE_HIGHT_THRESHOLD and suggested_nr <= 3:
                cell.fill = self.MODIFICARTION_FILL

        self._save()
        self._close()

    def _calc_received_num(self, order_nr_str, received_str):
        """
        计算实到数量。点货的时候可能在实到栏写加号，报单数量可能有欠货，在这里必须全部计算成实际数量。

        :param order_nr: 报单字符串
        :param received: 实到点货记录
        :return: 实际收到货物数量
        报单：10，      实到：+，    返回：10
        报单：欠5，      实到：+，    返回：5
        报单：欠5报5，    实到：+，   返回：10
        报单：xx，      实到：空，   返回：0
        """
        (ordered, owed)= utils._parse_order_string(order_nr_str)
        received = utils._parse_payed_received_string(received_str)
        if received is None:
            return 0
        if received == "OK":
            return ordered + owed
        # 如果执行到此处，received为实际收到数字
        return received


    def refresh_today_exceptions(self, good_op_log_file):
        e = self.calc_order_exceptions()

        self._open()
        exception_cn = self._get_column_cn("金额") + 2 # 金额后面第二列为异常插入位置
        inbound_num_cn = exception_cn + 1 # 入库数量列
        diff_num_cn = inbound_num_cn + 1 # 入库点货差
        self.ws.cell(row=1, column=exception_cn).value = "异常"
        self.ws.cell(row=1, column=inbound_num_cn).value = "入库数"
        self.ws.cell(row=1, column=diff_num_cn).value = "入库点货差"

        # 实际到货数量=快速上架数量+不绑定批次发货数量

        # 快速上架按编码汇总，注意sum()的返回结果是一个series，索引是商品编码
        df = pd.read_excel(good_op_log_file)
        tmp_df = df[df['操作类型'] == "快速上架"]
        fast_on_shelf_series = tmp_df.groupby("商品编码")['数量'].sum() # 快速上架数量汇总

        # 操作类型为“验货出库”，并且单据编号2==0的记录为不绑定批次发货记录
        tmp_df = df[(df['操作类型'] == "验货出库") & (df['单据编号2'] == 0)]
        no_bound_delivery_series = tmp_df.groupby("商品编码")['数量'].sum() # 不绑定批次发货数量汇总


        for i in range(2, self.ws.max_row + 1):
            p = str(self.ws.cell(row=i, column=self.provider_cn).value)
            good_code = str(self.ws.cell(row=i, column=self.good_code_cn).value)
            code = str(self.ws.cell(row=i, column=self.code_cn).value)
            spec = str(self.ws.cell(row=i, column=self.spec_cn).value)
            order_nr_str = utils.convert_possible_num_to_str(self.ws.cell(row=i, column=self.nr_cn).value) # 报单数量
            received_str = utils.convert_possible_num_to_str(self.ws.cell(row=i, column=self.received_cn).value) # 实到
            received = self._calc_received_num(order_nr_str, received_str)
            #print(received,order_nr_str,received_str)
            #v = "x"

            v1 = fast_on_shelf_series[good_code] if good_code in fast_on_shelf_series.index else 0
            v2 = no_bound_delivery_series[good_code] if good_code in no_bound_delivery_series.index else 0

            inbound_num = v1 + v2

            self.ws.cell(row=i, column=inbound_num_cn).value = inbound_num if inbound_num != 0 else "x"
            diff = inbound_num - received
            if diff != 0:
                self.ws.cell(row=i, column=diff_num_cn).value = diff

            if p not in e.keys():
                continue

            for l in e[p]:
                if code == l['code'] and spec == l['spec']:
                    self.ws.cell(row=i,column=exception_cn).value = utils.gen_text_for_one_exception_line(l,simplified=True)

        self._save()
        self._close()


if __name__ == "__main__":
    xp = XlsProcessor('./test.xlsx')
    orders = xp.gen_orders()
    import  utils
    print(utils.gen_all_orders_text(orders))

    up = [u'53d123', u'孙劲飞']
    print(xp.annotate_unknown_providers(up))
